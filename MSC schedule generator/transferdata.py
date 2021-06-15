import os
import shutil
import openpyxl as op

CONFIG_TABLE_ROW_START = 9
CONFIG_TABLE_COLUMN_START = 2
CONFIG_TABLE_ROW_END = 22
CONFIG_TABLE_COLUMN_END = 8
TUTOR_NAME_CELL = (3,3)
TUTOR_WORK_AWARD_CELL = (3,8)
TUTOR_DESIRED_HOURS_CELL = (5,8)

MSC_WORKERS_NEEDED_FILENAME = 'MSC Hours of Operation.xlsx'
AVAILABILITY_FOLDER_NAME = 'student availability'
AVAILABILITY_PROCESSED_FOLDER_NAME = 'student availability (already processed)'

MSC_CONSTRAINTS_FILENAME = 'MSC Tutor Constraints.xlsx'
MSC_TUTOR_SCHEDULE_FILENAME = 'MSC Tutor Schedule.xlsx'



files = os.listdir(AVAILABILITY_FOLDER_NAME)
for fn in files:
    if ".xlsx" in fn:
        nfn = "student availability (already processed)\\"+fn
        shutil.copy("Tutor Availability Form.xlsx",nfn)

        oldwb = op.open(AVAILABILITY_FOLDER_NAME+"\\"+fn)
        newwb = op.open(nfn)

        ows = oldwb.active
        nws = newwb.active

        for hour in range(CONFIG_TABLE_ROW_START,CONFIG_TABLE_ROW_END + 1):
            for day in range(CONFIG_TABLE_COLUMN_START, CONFIG_TABLE_COLUMN_END +1):
                nws.cell(hour,day).value = ows.cell(hour,day).value

        nws.cell(TUTOR_NAME_CELL[0],TUTOR_NAME_CELL[1]).value = ows.cell(TUTOR_NAME_CELL[0],TUTOR_NAME_CELL[1]).value
        nws.cell(TUTOR_DESIRED_HOURS_CELL[0],TUTOR_DESIRED_HOURS_CELL[1]).value = ows.cell(TUTOR_DESIRED_HOURS_CELL[0],TUTOR_DESIRED_HOURS_CELL[1]).value
        nws.cell(TUTOR_WORK_AWARD_CELL[0],TUTOR_WORK_AWARD_CELL[1]).value = ows.cell(TUTOR_WORK_AWARD_CELL[0],TUTOR_WORK_AWARD_CELL[1]).value        

        newwb.save(nfn)
        newwb.close()
        oldwb.close()
